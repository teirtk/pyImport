import io
import os
import operator
from datetime import timedelta, date
import pandas as pd
import numpy as np
import config
from openpyxl import load_workbook


def get_date(df):
    d = ""
    m = ""
    y = ""
    for row in df.itertuples():
        for (_, s) in enumerate(row):
            if isinstance(s, str):
                if s.startswith("Ngày"):
                    d = s[5:]
                    if len(d) == 1:
                        d = f"0{d}"
                    continue
                if s.startswith("Tháng"):
                    m = s[6:]
                    if len(m) == 1:
                        m = f"0{m}"
                    continue
                if s.startswith("Năm"):
                    y = s[4:]
                    if d <= "05":
                        d = "05"
                    elif d <= "15":
                        d = "15"
                    elif d <= "25":
                        d = "25"
                    else:
                        d = "05"
                        rdate = date.fromisoformat(f"{y}-{m}-{d}")
                        return rdate + timedelta(months=1)
                    return date.fromisoformat(f"{y}-{m}-{d}")
    return None


def do_process(file, conn):
    basename = os.path.basename(file)
    header = True
    buffer = io.StringIO()
    wb = load_workbook(file, data_only=True)
    try:
        for name in wb.sheetnames:
            if name == "KEYWORD" or wb[name]["C1"].value is None:
                continue
            ws = wb[name]
            df = pd.DataFrame(ws.values)
            fdate = get_date(df.head(7))
            df = df.iloc[7:, :14]
            df = df[df[0].notna()]
            df = df[df[1].notna()]
            df = df[df[0].str.contains(".", regex=False)]
            df = df.fillna(0)
            df.columns = ["cotID", "cotA", "cotB", "cotC", "cotD", "cotE", "cotF",
                          "cotG", "cotH", "cotI", "mota1", "mota2", "fdate", "fromFile"]
            tid = ws["C1"].value
            df["mota1"] = config.town_list[tid][4]
            df["mota2"] = config.town_list[tid][2]
            df["fdate"] = fdate
            df["fromFile"] = basename
            df.to_csv(buffer, index=False, header=header,
                      line_terminator="\n", sep=';')
            if header:
                header = False
        if buffer.getvalue().count('\n') > 1:
            buffer.seek(0)
            with conn.cursor() as cur:
                cur.execute(f"CREATE TEMP TABLE tmp_table ON COMMIT DROP AS "
                            f"TABLE {config.ext['caytrong']['schema']}.{config.ext['caytrong']['table']} WITH NO DATA;")
                cur.copy_expert(
                    "COPY tmp_table FROM STDIN WITH CSV HEADER DELIMITER as ';';", buffer)
                cur.execute(f"INSERT INTO {config.ext['caytrong']['schema']}.{config.ext['caytrong']['table']} "
                            f"SELECT * FROM tmp_table EXCEPT "
                            f"SELECT * FROM {config.ext['caytrong']['schema']}.{config.ext['caytrong']['table']};")
                nline = cur.rowcount
                conn.commit()
            if nline:
                return f"{basename}: {nline:,} dòng được thêm \n"
            else:
                return f"{basename}: Dữ liệu đã có (bỏ qua) \n"
    except TypeError:
        return f"{basename} bị lỗi\n"
    finally:
        buffer.close()
    return ""
